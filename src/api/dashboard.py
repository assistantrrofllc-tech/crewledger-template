"""
Dashboard routes — serves the web UI and receipt images.

All management views: home, ledger, employee management, settings, projects.
Receipt images served from local storage with path traversal protection.
Export endpoints: QuickBooks CSV, Google Sheets CSV, Excel (.xlsx).
Receipt editing with audit trail. Notes support throughout.
"""

import csv
import io
import json
import logging
import secrets
from datetime import datetime, timedelta
from pathlib import Path

from flask import (
    Blueprint, render_template, send_from_directory, jsonify, request, abort,
    Response, send_file,
)

from config.settings import RECEIPT_STORAGE_PATH, CERT_STORAGE_PATH
from src.database.connection import get_db
from src.services.cert_status import calculate_cert_status, days_until_expiry
from src.services.permissions import check_permission

log = logging.getLogger(__name__)

dashboard_bp = Blueprint("dashboard", __name__)

# Per-module sub-navigation (Layer 2)
MODULE_NAVS = {
    "crewledger": [
        {"id": "dashboard", "label": "Dashboard", "href": "/"},
        {"id": "ledger", "label": "Ledger", "href": "/ledger"},
        {"id": "projects", "label": "Projects", "href": "/projects"},
        {"id": "settings", "label": "Settings", "href": "/settings"},
    ],
    "crewcert": [
        {"id": "dashboard", "label": "Dashboard", "href": "/crewcert"},
        {"id": "employees", "label": "Employees", "href": "/crew"},
    ],
}


def _render_module(template, active_module, active_subnav="", **kwargs):
    """Render a template with module navigation context."""
    return render_template(
        template,
        active_module=active_module,
        active_subnav=active_subnav,
        module_nav=MODULE_NAVS.get(active_module, []),
        **kwargs,
    )


# ── Pages ────────────────────────────────────────────────────


@dashboard_bp.route("/")
def home():
    """Dashboard home — spend summary, flagged receipts, recent activity."""
    db = get_db()
    try:
        stats = _get_dashboard_stats(db)
        flagged = _get_flagged_receipts(db)
        recent = _get_recent_receipts(db, limit=10)
        unknown = _get_unknown_contacts(db, limit=10)
        can_edit = check_permission(None, "crewledger", "edit")
        return _render_module("index.html", "crewledger", "dashboard", stats=stats, flagged=flagged, recent=recent, unknown=unknown, can_edit=can_edit)
    finally:
        db.close()


# ── Receipt Image Serving ────────────────────────────────────


@dashboard_bp.route("/receipts/image/<filename>")
def serve_receipt_image(filename):
    """Serve a receipt image from local storage.

    Path traversal protection: only serves files from the receipts directory,
    filename must not contain path separators.
    """
    # Block path traversal
    if "/" in filename or "\\" in filename or ".." in filename:
        abort(404)

    storage_dir = Path(RECEIPT_STORAGE_PATH).resolve()
    file_path = (storage_dir / filename).resolve()

    # Ensure the resolved path is inside the storage directory
    if not str(file_path).startswith(str(storage_dir)):
        abort(404)

    if not file_path.exists():
        abort(404)

    return send_from_directory(str(storage_dir), filename)


# ── Cert Document Serving ────────────────────────────────────


@dashboard_bp.route("/certs/file/<filename>")
def serve_cert_file(filename):
    """Serve a cert PDF from the cert_files directory.

    Path traversal protection: filename must not contain path separators.
    Files stored at storage/certifications/cert_files/<filename>.
    """
    if "/" in filename or "\\" in filename or ".." in filename:
        abort(404)

    storage_dir = (Path(CERT_STORAGE_PATH) / "cert_files").resolve()
    file_path = (storage_dir / filename).resolve()

    if not str(file_path).startswith(str(storage_dir)):
        abort(404)
    if not file_path.exists():
        abort(404)

    return send_from_directory(str(storage_dir), filename, mimetype="application/pdf")


@dashboard_bp.route("/certifications/document/<employee_uuid>/<filename>")
def serve_cert_document(employee_uuid, filename):
    """Serve a cert document from local storage.

    Path traversal protection: validates UUID format and filename.
    Files stored at storage/certifications/<employee_uuid>/<filename>.
    """
    # Block path traversal
    if "/" in filename or "\\" in filename or ".." in filename:
        abort(404)
    if ".." in employee_uuid or "/" in employee_uuid or "\\" in employee_uuid:
        abort(404)

    storage_dir = Path(CERT_STORAGE_PATH).resolve() / employee_uuid
    file_path = (storage_dir / filename).resolve()

    # Ensure resolved path is inside the cert storage directory
    if not str(file_path).startswith(str(Path(CERT_STORAGE_PATH).resolve())):
        abort(404)

    if not file_path.exists():
        abort(404)

    return send_from_directory(str(storage_dir), filename)


# ── API Endpoints ────────────────────────────────────────────


@dashboard_bp.route("/api/receipts")
def api_receipts():
    """JSON endpoint for receipt data with filtering.

    Query params:
        period: today, week, month, ytd, all (default: all)
        start: YYYY-MM-DD custom start date
        end: YYYY-MM-DD custom end date
        employee: employee ID
        project: project ID
        vendor: vendor name (partial match)
        status: confirmed, pending, flagged
        sort: date, employee, vendor, project, amount, status (default: date)
        order: asc, desc (default: desc)
    """
    db = get_db()
    try:
        receipts = _query_receipts(db, request.args)
        return jsonify(receipts)
    finally:
        db.close()


@dashboard_bp.route("/api/receipts", methods=["POST"])
def api_create_receipt():
    """Manually create a receipt (management entry). Saved as confirmed."""
    data = request.get_json(silent=True) or {}

    employee_id = data.get("employee_id")
    vendor_name = data.get("vendor_name", "").strip()
    total = data.get("total")

    if not employee_id:
        return jsonify({"error": "Employee is required"}), 400
    if not vendor_name:
        return jsonify({"error": "Vendor name is required"}), 400
    if not total or float(total) <= 0:
        return jsonify({"error": "A valid total is required"}), 400

    db = get_db()
    try:
        emp = db.execute("SELECT id FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not emp:
            return jsonify({"error": "Employee not found"}), 404

        project_id = data.get("project_id")
        if project_id:
            proj = db.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone()
            if not proj:
                return jsonify({"error": "Project not found"}), 404

        category_id = data.get("category_id") or None

        cursor = db.execute(
            """INSERT INTO receipts
               (employee_id, project_id, category_id, vendor_name, purchase_date, subtotal, tax, total,
                payment_method, notes, status, confirmed_at, is_missed_receipt)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'confirmed', datetime('now'), 1)""",
            (
                employee_id,
                project_id,
                category_id,
                vendor_name,
                data.get("purchase_date"),
                data.get("subtotal") or 0,
                data.get("tax") or 0,
                float(total),
                data.get("payment_method", ""),
                data.get("notes", ""),
            ),
        )
        receipt_id = cursor.lastrowid
        db.commit()

        log.info("Manual receipt #%d created by management (vendor=%s, total=%s)", receipt_id, vendor_name, total)
        return jsonify({"status": "created", "id": receipt_id}), 201
    finally:
        db.close()


@dashboard_bp.route("/api/receipts/export")
def api_receipts_export():
    """Export filtered receipts as QuickBooks CSV, Google Sheets CSV, or Excel.

    Uses the same filters as the main receipts API.
    Query param 'format': quickbooks, csv, excel (default: csv)
    """
    db = get_db()
    try:
        receipts = _query_receipts(db, request.args)
        fmt = request.args.get("format", "csv")

        if fmt == "quickbooks":
            return _export_quickbooks_csv(receipts)
        elif fmt == "excel":
            return _export_excel(receipts)
        else:
            return _export_csv(receipts)
    finally:
        db.close()


@dashboard_bp.route("/api/receipts/<int:receipt_id>")
def api_receipt_detail(receipt_id):
    """Single receipt with full detail including line items."""
    db = get_db()
    try:
        receipt = _get_receipt_detail(db, receipt_id)
        if not receipt:
            return jsonify({"error": "Receipt not found"}), 404
        return jsonify(receipt)
    finally:
        db.close()


@dashboard_bp.route("/api/dashboard/stats")
def api_dashboard_stats():
    """Dashboard summary stats as JSON."""
    db = get_db()
    try:
        return jsonify(_get_dashboard_stats(db))
    finally:
        db.close()


# ── Employee Management ──────────────────────────────────


@dashboard_bp.route("/employees")
def employees_page():
    """Employee management page."""
    db = get_db()
    try:
        employees = db.execute("""
            SELECT e.*,
                   (SELECT MAX(r.created_at) FROM receipts r WHERE r.employee_id = e.id) as last_submission
            FROM employees e ORDER BY e.first_name
        """).fetchall()
        return _render_module("employees.html", "crewledger", "", employees=[dict(e) for e in employees])
    finally:
        db.close()


@dashboard_bp.route("/api/employees", methods=["GET"])
def api_employees():
    """List all employees as JSON."""
    db = get_db()
    try:
        rows = db.execute("""
            SELECT e.*,
                   (SELECT MAX(r.created_at) FROM receipts r WHERE r.employee_id = e.id) as last_submission
            FROM employees e ORDER BY e.first_name
        """).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        db.close()


@dashboard_bp.route("/api/employees", methods=["POST"])
def api_add_employee():
    """Add a new employee."""
    data = request.get_json()
    if not data or not data.get("first_name") or not data.get("phone_number"):
        return jsonify({"error": "first_name and phone_number are required"}), 400

    phone = data["phone_number"].strip()
    if not phone.startswith("+"):
        phone = "+1" + phone.replace("-", "").replace(" ", "").replace("(", "").replace(")", "")

    db = get_db()
    try:
        existing = db.execute("SELECT id FROM employees WHERE phone_number = ?", (phone,)).fetchone()
        if existing:
            return jsonify({"error": "Phone number already registered"}), 409

        token = secrets.token_urlsafe(12)
        db.execute(
            "INSERT INTO employees (phone_number, first_name, full_name, email, role, crew, public_token) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (phone, data["first_name"], data.get("full_name"), data.get("email"), data.get("role"), data.get("crew"), token),
        )
        db.commit()
        return jsonify({"status": "created", "phone_number": phone}), 201
    finally:
        db.close()


@dashboard_bp.route("/api/employees/<int:employee_id>", methods=["GET"])
def api_employee_detail(employee_id):
    """Get a single employee (also serves as CrewCert QR landing page)."""
    db = get_db()
    try:
        row = db.execute("SELECT * FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not row:
            return jsonify({"error": "Employee not found"}), 404
        return jsonify(dict(row))
    finally:
        db.close()


@dashboard_bp.route("/api/employees/<int:employee_id>", methods=["PUT"])
def api_update_employee(employee_id):
    """Update employee fields."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    allowed = {"first_name", "full_name", "email", "role", "crew", "phone_number", "notes", "nickname", "is_driver"}
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({"error": "No valid fields to update"}), 400

    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [employee_id]

    db = get_db()
    try:
        db.execute(f"UPDATE employees SET {set_clause}, updated_at = datetime('now') WHERE id = ?", values)
        db.commit()
        return jsonify({"status": "updated"})
    finally:
        db.close()


@dashboard_bp.route("/api/employees/<int:employee_id>/deactivate", methods=["POST"])
def api_deactivate_employee(employee_id):
    """Deactivate an employee — they can no longer submit receipts."""
    db = get_db()
    try:
        db.execute("UPDATE employees SET is_active = 0, updated_at = datetime('now') WHERE id = ?", (employee_id,))
        db.commit()
        return jsonify({"status": "deactivated"})
    finally:
        db.close()


@dashboard_bp.route("/api/employees/<int:employee_id>/activate", methods=["POST"])
def api_activate_employee(employee_id):
    """Reactivate an employee."""
    db = get_db()
    try:
        db.execute("UPDATE employees SET is_active = 1, updated_at = datetime('now') WHERE id = ?", (employee_id,))
        db.commit()
        return jsonify({"status": "activated"})
    finally:
        db.close()


# ── Project Management ────────────────────────────────────


@dashboard_bp.route("/api/projects", methods=["GET"])
def api_projects():
    """List all projects as JSON."""
    db = get_db()
    try:
        rows = db.execute("""
            SELECT p.*,
                   (SELECT COUNT(*) FROM receipts r
                    WHERE r.status NOT IN ('deleted', 'duplicate')
                      AND (r.project_id = p.id OR r.matched_project_name = p.name)
                   ) as receipt_count,
                   (SELECT COALESCE(SUM(r.total), 0) FROM receipts r
                    WHERE r.status NOT IN ('deleted', 'duplicate')
                      AND (r.project_id = p.id OR r.matched_project_name = p.name)
                   ) as total_spend
            FROM projects p ORDER BY p.name
        """).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        db.close()


@dashboard_bp.route("/api/projects", methods=["POST"])
def api_add_project():
    """Add a new project."""
    data = request.get_json()
    if not data or not data.get("name"):
        return jsonify({"error": "Project name is required"}), 400

    db = get_db()
    try:
        existing = db.execute("SELECT id FROM projects WHERE name = ?", (data["name"],)).fetchone()
        if existing:
            return jsonify({"error": "Project name already exists"}), 409

        db.execute(
            """INSERT INTO projects (project_code, name, address, city, state, status, start_date, end_date, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                data.get("project_code"),
                data["name"],
                data.get("address"),
                data.get("city"),
                data.get("state"),
                data.get("status", "active"),
                data.get("start_date"),
                data.get("end_date"),
                data.get("notes"),
            ),
        )
        db.commit()
        return jsonify({"status": "created", "name": data["name"]}), 201
    finally:
        db.close()


@dashboard_bp.route("/api/projects/<int:project_id>", methods=["PUT"])
def api_update_project(project_id):
    """Update a project."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    allowed = {"project_code", "name", "address", "city", "state", "status", "start_date", "end_date", "notes"}
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({"error": "No valid fields to update"}), 400

    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [project_id]

    db = get_db()
    try:
        db.execute(f"UPDATE projects SET {set_clause}, updated_at = datetime('now') WHERE id = ?", values)
        db.commit()
        return jsonify({"status": "updated"})
    finally:
        db.close()


@dashboard_bp.route("/api/projects/<int:project_id>", methods=["DELETE"])
def api_delete_project(project_id):
    """Delete a project. Receipts linked to it are kept but unlinked."""
    db = get_db()
    try:
        row = db.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not row:
            return jsonify({"error": "Project not found"}), 404
        db.execute("UPDATE receipts SET project_id = NULL WHERE project_id = ?", (project_id,))
        db.execute("DELETE FROM projects WHERE id = ?", (project_id,))
        db.commit()
        return jsonify({"status": "deleted"})
    finally:
        db.close()


@dashboard_bp.route("/api/projects/<int:project_id>", methods=["GET"])
def api_project_detail(project_id):
    """Get a single project."""
    db = get_db()
    try:
        row = db.execute("""
            SELECT p.*,
                   (SELECT COUNT(*) FROM receipts r WHERE r.project_id = p.id) as receipt_count,
                   (SELECT COALESCE(SUM(r.total), 0) FROM receipts r WHERE r.project_id = p.id) as total_spend
            FROM projects p WHERE p.id = ?
        """, (project_id,)).fetchone()
        if not row:
            return jsonify({"error": "Project not found"}), 404
        return jsonify(dict(row))
    finally:
        db.close()


# ── Categories ───────────────────────────────────────────


@dashboard_bp.route("/api/categories", methods=["GET"])
def api_categories():
    """List all categories. Pass ?active=1 to get only active ones."""
    db = get_db()
    try:
        active_only = request.args.get("active", "0")
        if active_only == "1":
            rows = db.execute("SELECT * FROM categories WHERE is_active = 1 ORDER BY sort_order, name").fetchall()
        else:
            rows = db.execute("SELECT * FROM categories ORDER BY sort_order, name").fetchall()
        # Include receipt count for management UI
        result = []
        for r in rows:
            d = dict(r)
            count = db.execute(
                "SELECT COUNT(*) as cnt FROM receipts WHERE category_id = ? AND status NOT IN ('deleted','duplicate')",
                (r["id"],),
            ).fetchone()
            d["receipt_count"] = count["cnt"] if count else 0
            result.append(d)
        return jsonify(result)
    finally:
        db.close()


@dashboard_bp.route("/api/categories", methods=["POST"])
def api_add_category():
    """Add a new category."""
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Name is required"}), 400

    db = get_db()
    try:
        existing = db.execute("SELECT id FROM categories WHERE LOWER(name) = LOWER(?)", (name,)).fetchone()
        if existing:
            return jsonify({"error": "Category already exists"}), 409
        max_order = db.execute("SELECT MAX(sort_order) as m FROM categories").fetchone()["m"] or 0
        cursor = db.execute(
            "INSERT INTO categories (name, description, sort_order) VALUES (?, ?, ?)",
            (name, data.get("description", ""), max_order + 1),
        )
        db.commit()
        return jsonify({"status": "created", "id": cursor.lastrowid, "name": name}), 201
    finally:
        db.close()


@dashboard_bp.route("/api/categories/<int:cat_id>", methods=["PUT"])
def api_update_category(cat_id):
    """Rename or update a category."""
    data = request.get_json(silent=True) or {}
    db = get_db()
    try:
        cat = db.execute("SELECT * FROM categories WHERE id = ?", (cat_id,)).fetchone()
        if not cat:
            return jsonify({"error": "Category not found"}), 404

        name = (data.get("name") or "").strip()
        if name and name != cat["name"]:
            dup = db.execute("SELECT id FROM categories WHERE LOWER(name) = LOWER(?) AND id != ?", (name, cat_id)).fetchone()
            if dup:
                return jsonify({"error": "A category with that name already exists"}), 409
            # Count receipts using this category for the warning
            count = db.execute(
                "SELECT COUNT(*) as cnt FROM receipts WHERE category_id = ?", (cat_id,)
            ).fetchone()["cnt"]
            db.execute("UPDATE categories SET name = ? WHERE id = ?", (name, cat_id))
            db.commit()
            return jsonify({"status": "updated", "receipt_count": count})

        return jsonify({"status": "no_change"})
    finally:
        db.close()


@dashboard_bp.route("/api/categories/<int:cat_id>/deactivate", methods=["POST"])
def api_deactivate_category(cat_id):
    """Deactivate a category — hidden from dropdowns, historical receipts keep it."""
    db = get_db()
    try:
        cat = db.execute("SELECT * FROM categories WHERE id = ?", (cat_id,)).fetchone()
        if not cat:
            return jsonify({"error": "Category not found"}), 404
        db.execute("UPDATE categories SET is_active = 0 WHERE id = ?", (cat_id,))
        db.commit()
        return jsonify({"status": "deactivated"})
    finally:
        db.close()


@dashboard_bp.route("/api/categories/<int:cat_id>/activate", methods=["POST"])
def api_activate_category(cat_id):
    """Reactivate a deactivated category."""
    db = get_db()
    try:
        cat = db.execute("SELECT * FROM categories WHERE id = ?", (cat_id,)).fetchone()
        if not cat:
            return jsonify({"error": "Category not found"}), 404
        db.execute("UPDATE categories SET is_active = 1 WHERE id = ?", (cat_id,))
        db.commit()
        return jsonify({"status": "activated"})
    finally:
        db.close()


# ── Unknown Contacts ─────────────────────────────────────


@dashboard_bp.route("/api/unknown-contacts")
def api_unknown_contacts():
    """List recent unknown contact attempts."""
    db = get_db()
    try:
        rows = db.execute("""
            SELECT * FROM unknown_contacts ORDER BY created_at DESC LIMIT 50
        """).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        db.close()


# ── Ledger Page ──────────────────────────────────────────


@dashboard_bp.route("/ledger")
def ledger_page():
    """Banking-style transaction ledger."""
    db = get_db()
    try:
        employees = db.execute("SELECT id, first_name FROM employees ORDER BY first_name").fetchall()
        projects = db.execute("SELECT id, name FROM projects WHERE status = 'active' ORDER BY name").fetchall()
        categories = db.execute("SELECT id, name FROM categories ORDER BY name").fetchall()
        can_edit = check_permission(None, "crewledger", "edit")
        return _render_module(
            "ledger.html", "crewledger", "ledger",
            employees=[dict(e) for e in employees],
            projects=[dict(p) for p in projects],
            categories=[dict(c) for c in categories],
            can_edit=can_edit,
        )
    finally:
        db.close()


@dashboard_bp.route("/crewcert")
def crewcert_home():
    """CrewCert module home — certification dashboard with summary and alerts."""
    if not check_permission(None, "crewcert", "view"):
        abort(403)
    return _render_module("crewcert_dashboard.html", "crewcert", "dashboard")


@dashboard_bp.route("/crew")
def crew_page():
    """CrewCert module — employee roster and certification tracking."""
    if not check_permission(None, "crewcert", "view"):
        abort(403)
    return _render_module("crew.html", "crewcert", "employees")


@dashboard_bp.route("/crew/<int:employee_id>")
def crew_detail_page(employee_id):
    """CrewCert — individual employee profile and certifications."""
    db = get_db()
    try:
        emp = db.execute("SELECT * FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not emp:
            abort(404)

        cert_types = db.execute(
            "SELECT * FROM certification_types WHERE is_active = 1 ORDER BY sort_order"
        ).fetchall()

        return _render_module(
            "crew_detail.html", "crewcert", "employees",
            employee=dict(emp),
            cert_types=[dict(ct) for ct in cert_types],
        )
    finally:
        db.close()


# ── Public Cert Verification ─────────────────────────────


# Simple in-memory rate limiter: token -> list of timestamps
_scan_rate_limit: dict[str, list[float]] = {}


@dashboard_bp.route("/crew/verify/<token>")
def public_verify(token):
    """Public cert verification page — no login required.

    Looked up by public_token. Shows employee name + certs only.
    Rate limited to 30 requests per token per hour.
    """
    import time

    # Rate limiting: 30 requests per token per hour
    now = time.time()
    window = _scan_rate_limit.get(token, [])
    window = [t for t in window if now - t < 3600]
    if len(window) >= 30:
        return render_template("verify_rate_limited.html"), 429
    window.append(now)
    _scan_rate_limit[token] = window

    db = get_db()
    try:
        emp = db.execute(
            "SELECT id, first_name, full_name, photo, is_active, public_token FROM employees WHERE public_token = ?",
            (token,),
        ).fetchone()

        if not emp:
            return render_template("verify_invalid.html"), 404

        if not emp["is_active"]:
            return render_template("verify_inactive.html"), 200

        # Log the scan
        ip = request.headers.get("X-Forwarded-For", request.remote_addr)
        if ip and "," in ip:
            ip = ip.split(",")[0].strip()
        ua = request.headers.get("User-Agent", "")[:200]
        db.execute(
            "INSERT INTO qr_scan_log (employee_id, ip_address, user_agent) VALUES (?, ?, ?)",
            (emp["id"], ip, ua),
        )
        db.commit()

        # Get active certs
        certs = db.execute("""
            SELECT c.id, ct.name as cert_name, c.issued_at, c.expires_at, c.document_path
            FROM certifications c
            JOIN certification_types ct ON c.cert_type_id = ct.id
            WHERE c.employee_id = ? AND c.is_active = 1
            ORDER BY ct.sort_order
        """, (emp["id"],)).fetchall()

        cert_list = []
        for c in certs:
            status = calculate_cert_status(c["expires_at"])
            has_doc = bool(c["document_path"])
            if has_doc:
                doc_full = Path(CERT_STORAGE_PATH) / c["document_path"]
                has_doc = doc_full.exists()
            cert_list.append({
                "id": c["id"],
                "name": c["cert_name"],
                "issued_at": c["issued_at"],
                "expires_at": c["expires_at"],
                "status": status,
                "has_document": has_doc,
            })

        return render_template(
            "verify_public.html",
            employee_name=emp["full_name"] or emp["first_name"],
            employee_photo=emp["photo"],
            certs=cert_list,
            token=token,
            verified_at=datetime.now().strftime("%b %d, %Y %I:%M%p"),
        )
    finally:
        db.close()


@dashboard_bp.route("/crew/verify/<token>/cert/<int:cert_id>")
def public_verify_cert(token, cert_id):
    """Serve a cert document publicly — scoped to the employee's token.

    No login required. Token must belong to an active employee.
    cert_id must belong to that employee. Rate limited same as verify page.
    """
    import time

    # Rate limiting: shared with verify page (30 req/hr/token)
    now = time.time()
    window = _scan_rate_limit.get(token, [])
    window = [t for t in window if now - t < 3600]
    if len(window) >= 30:
        return render_template("verify_rate_limited.html"), 429
    window.append(now)
    _scan_rate_limit[token] = window

    db = get_db()
    try:
        emp = db.execute(
            "SELECT id, is_active, employee_uuid FROM employees WHERE public_token = ?",
            (token,),
        ).fetchone()

        if not emp:
            return render_template("verify_invalid.html"), 404
        if not emp["is_active"]:
            return render_template("verify_inactive.html"), 200

        # Cert must belong to this employee
        cert = db.execute(
            "SELECT id, document_path FROM certifications WHERE id = ? AND employee_id = ? AND is_active = 1",
            (cert_id, emp["id"]),
        ).fetchone()

        if not cert:
            abort(404)

        # Log the document view
        ip = request.headers.get("X-Forwarded-For", request.remote_addr)
        if ip and "," in ip:
            ip = ip.split(",")[0].strip()
        ua = request.headers.get("User-Agent", "")[:200]
        db.execute(
            "INSERT INTO qr_scan_log (employee_id, ip_address, user_agent) VALUES (?, ?, ?)",
            (emp["id"], ip, ua),
        )
        db.commit()

        # Check document exists
        if not cert["document_path"]:
            return render_template("verify_no_document.html"), 200

        doc_path = (Path(CERT_STORAGE_PATH) / cert["document_path"]).resolve()
        if not str(doc_path).startswith(str(Path(CERT_STORAGE_PATH).resolve())):
            abort(404)
        if not doc_path.exists():
            return render_template("verify_no_document.html"), 200

        return send_file(doc_path, as_attachment=False)
    finally:
        db.close()


@dashboard_bp.route("/api/crew/employees/<int:employee_id>/qr")
def api_employee_qr(employee_id):
    """Generate QR code PNG for an employee's public verify URL."""
    import qrcode
    from io import BytesIO

    db = get_db()
    try:
        emp = db.execute("SELECT public_token FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not emp or not emp["public_token"]:
            return jsonify({"error": "Employee not found or no token"}), 404

        host = request.host_url.rstrip("/")
        url = f"{host}/crew/verify/{emp['public_token']}"

        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=4)
        qr.add_data(url)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")

        buf = BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)

        return send_file(buf, mimetype="image/png", as_attachment=False, download_name=f"qr_{employee_id}.png")
    finally:
        db.close()


@dashboard_bp.route("/api/crew/employees/<int:employee_id>/regenerate-token", methods=["POST"])
def api_regenerate_token(employee_id):
    """Regenerate public_token for an employee, invalidating old QR code."""
    db = get_db()
    try:
        emp = db.execute("SELECT id FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not emp:
            return jsonify({"error": "Employee not found"}), 404

        new_token = secrets.token_urlsafe(12)
        db.execute(
            "UPDATE employees SET public_token = ?, updated_at = datetime('now') WHERE id = ?",
            (new_token, employee_id),
        )
        db.commit()
        return jsonify({"status": "regenerated", "token": new_token})
    finally:
        db.close()


@dashboard_bp.route("/api/crew/employees/<int:employee_id>/scan-log")
def api_employee_scan_log(employee_id):
    """Return the last 20 QR scans for an employee."""
    db = get_db()
    try:
        rows = db.execute("""
            SELECT scanned_at, ip_address
            FROM qr_scan_log
            WHERE employee_id = ?
            ORDER BY scanned_at DESC
            LIMIT 20
        """, (employee_id,)).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        db.close()


# ── CrewCert Dashboard API ───────────────────────────────


@dashboard_bp.route("/api/crewcert/dashboard")
def api_crewcert_dashboard():
    """Dashboard summary: counts, active alerts, upcoming expirations."""
    db = get_db()
    try:
        # Summary counts
        total_employees = db.execute(
            "SELECT COUNT(*) as cnt FROM employees WHERE is_active = 1"
        ).fetchone()["cnt"]

        # Get all active certs with expiry info
        certs = db.execute("""
            SELECT c.id, c.employee_id, c.expires_at, c.cert_type_id,
                   ct.name as cert_type_name,
                   e.first_name, e.full_name
            FROM certifications c
            JOIN certification_types ct ON c.cert_type_id = ct.id
            JOIN employees e ON c.employee_id = e.id
            WHERE c.is_active = 1 AND e.is_active = 1
        """).fetchall()

        expired_count = 0
        expiring_count = 0
        no_expiry_count = 0
        for c in certs:
            status = calculate_cert_status(c["expires_at"])
            if status == "expired":
                expired_count += 1
            elif status == "expiring":
                expiring_count += 1
            elif status == "no_expiry":
                no_expiry_count += 1

        # Active (unacknowledged) alerts
        alerts = db.execute("""
            SELECT ca.*, e.first_name, e.full_name,
                   ct.name as cert_type_name, c.expires_at
            FROM cert_alerts ca
            JOIN employees e ON ca.employee_id = e.id
            JOIN certifications c ON ca.cert_id = c.id
            JOIN certification_types ct ON c.cert_type_id = ct.id
            WHERE ca.acknowledged = 0
            ORDER BY
                CASE ca.alert_type WHEN 'expired' THEN 0 WHEN 'expiring' THEN 1 ELSE 2 END,
                ca.created_at DESC
        """).fetchall()

        alert_list = []
        for a in alerts:
            days = days_until_expiry(a["expires_at"])
            alert_list.append({
                "id": a["id"],
                "employee_id": a["employee_id"],
                "cert_id": a["cert_id"],
                "employee_name": a["full_name"] or a["first_name"],
                "cert_name": a["cert_type_name"],
                "alert_type": a["alert_type"],
                "days_until_expiry": days,
                "expires_at": a["expires_at"],
                "created_at": a["created_at"],
            })

        # Upcoming expirations (next 90 days) for calendar
        upcoming = db.execute("""
            SELECT c.expires_at, ct.name as cert_type_name,
                   e.first_name, e.full_name, e.id as employee_id
            FROM certifications c
            JOIN certification_types ct ON c.cert_type_id = ct.id
            JOIN employees e ON c.employee_id = e.id
            WHERE c.is_active = 1 AND e.is_active = 1
              AND c.expires_at IS NOT NULL AND c.expires_at != ''
              AND c.expires_at >= date('now')
              AND c.expires_at <= date('now', '+90 days')
            ORDER BY c.expires_at
        """).fetchall()

        calendar = []
        for u in upcoming:
            calendar.append({
                "date": u["expires_at"],
                "cert_name": u["cert_type_name"],
                "employee_name": u["full_name"] or u["first_name"],
                "employee_id": u["employee_id"],
            })

        return jsonify({
            "total_employees": total_employees,
            "expired_count": expired_count,
            "expiring_count": expiring_count,
            "no_expiry_count": no_expiry_count,
            "alerts": alert_list,
            "calendar": calendar,
        })
    finally:
        db.close()


@dashboard_bp.route("/api/crewcert/alerts/<int:alert_id>/acknowledge", methods=["POST"])
def api_acknowledge_alert(alert_id):
    """Acknowledge (dismiss) a cert alert."""
    db = get_db()
    try:
        alert = db.execute("SELECT id FROM cert_alerts WHERE id = ?", (alert_id,)).fetchone()
        if not alert:
            return jsonify({"error": "Alert not found"}), 404
        db.execute(
            "UPDATE cert_alerts SET acknowledged = 1, acknowledged_at = datetime('now'), acknowledged_by = 'dashboard' WHERE id = ?",
            (alert_id,),
        )
        db.commit()
        return jsonify({"status": "acknowledged"})
    finally:
        db.close()


@dashboard_bp.route("/api/crewcert/refresh", methods=["POST"])
def api_refresh_cert_status():
    """Manually trigger a cert status refresh."""
    from src.services.cert_refresh import run_cert_status_refresh
    result = run_cert_status_refresh()
    return jsonify({"status": "refreshed", **result})


# ── CrewCert API ─────────────────────────────────────────


@dashboard_bp.route("/api/cert-types")
def api_cert_types():
    """List all certification types."""
    db = get_db()
    try:
        rows = db.execute(
            "SELECT * FROM certification_types WHERE is_active = 1 ORDER BY sort_order"
        ).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        db.close()


@dashboard_bp.route("/api/crew/employees")
def api_crew_employees():
    """Employee roster with certification badge summary.

    Returns each employee with a `certs` array showing the status
    of every cert type (valid / expiring / expired / none).
    """
    db = get_db()
    try:
        cert_types = db.execute(
            "SELECT id, name, slug FROM certification_types WHERE is_active = 1 ORDER BY sort_order"
        ).fetchall()

        employees = db.execute("""
            SELECT id, employee_uuid, first_name, full_name, phone_number,
                   email, role, crew, is_active, nickname, is_driver
            FROM employees ORDER BY first_name
        """).fetchall()

        result = []
        for emp in employees:
            emp_dict = dict(emp)

            # Get this employee's most recent cert per type
            certs = db.execute("""
                SELECT c.cert_type_id, c.issued_at, c.expires_at
                FROM certifications c
                WHERE c.employee_id = ? AND c.is_active = 1
                ORDER BY c.issued_at DESC
            """, (emp["id"],)).fetchall()

            # Build a lookup: cert_type_id -> most recent cert
            cert_lookup = {}
            for c in certs:
                if c["cert_type_id"] not in cert_lookup:
                    cert_lookup[c["cert_type_id"]] = dict(c)

            # Build badge array
            badges = []
            has_expired = False
            has_expiring = False
            for ct in cert_types:
                cert = cert_lookup.get(ct["id"])
                if cert:
                    status = calculate_cert_status(cert["expires_at"])
                else:
                    status = "none"

                if status == "expired":
                    has_expired = True
                elif status == "expiring":
                    has_expiring = True

                badges.append({
                    "type_id": ct["id"],
                    "slug": ct["slug"],
                    "name": ct["name"],
                    "status": status,
                })

            emp_dict["certs"] = badges
            emp_dict["has_expired"] = has_expired
            emp_dict["has_expiring"] = has_expiring
            result.append(emp_dict)

        return jsonify(result)
    finally:
        db.close()


@dashboard_bp.route("/api/crew/employees/<int:employee_id>/certs")
def api_employee_certs(employee_id):
    """Full certification list for one employee."""
    db = get_db()
    try:
        emp = db.execute("SELECT id FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not emp:
            return jsonify({"error": "Employee not found"}), 404

        rows = db.execute("""
            SELECT c.*, ct.name as cert_type_name, ct.slug as cert_type_slug
            FROM certifications c
            JOIN certification_types ct ON c.cert_type_id = ct.id
            WHERE c.employee_id = ? AND c.is_active = 1
            ORDER BY ct.sort_order
        """, (employee_id,)).fetchall()

        certs = []
        for r in rows:
            d = dict(r)
            d["status"] = calculate_cert_status(r["expires_at"])
            certs.append(d)

        return jsonify(certs)
    finally:
        db.close()


@dashboard_bp.route("/api/crew/certifications", methods=["POST"])
def api_add_certification():
    """Add a new certification record."""
    data = request.get_json(silent=True) or {}
    employee_id = data.get("employee_id")
    cert_type_id = data.get("cert_type_id")

    if not employee_id or not cert_type_id:
        return jsonify({"error": "employee_id and cert_type_id are required"}), 400

    db = get_db()
    try:
        cursor = db.execute(
            """INSERT INTO certifications
               (employee_id, cert_type_id, issued_at, expires_at, issuing_org, notes)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (
                employee_id, cert_type_id,
                data.get("issued_at"), data.get("expires_at"),
                data.get("issuing_org"), data.get("notes"),
            ),
        )
        db.commit()
        return jsonify({"status": "created", "id": cursor.lastrowid}), 201
    except Exception as e:
        if "UNIQUE constraint" in str(e):
            return jsonify({"error": "Duplicate certification record"}), 409
        raise
    finally:
        db.close()


@dashboard_bp.route("/api/crew/certifications/<int:cert_id>", methods=["PUT"])
def api_update_certification(cert_id):
    """Update a certification record."""
    data = request.get_json(silent=True) or {}
    db = get_db()
    try:
        cert = db.execute("SELECT id FROM certifications WHERE id = ? AND is_active = 1", (cert_id,)).fetchone()
        if not cert:
            return jsonify({"error": "Certification not found"}), 404

        allowed = {"cert_type_id", "issued_at", "expires_at", "issuing_org", "notes", "document_path"}
        updates = {k: v for k, v in data.items() if k in allowed}
        if not updates:
            return jsonify({"error": "No valid fields to update"}), 400

        set_clause = ", ".join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [cert_id]
        db.execute(f"UPDATE certifications SET {set_clause}, updated_at = datetime('now') WHERE id = ?", values)
        db.commit()
        return jsonify({"status": "updated"})
    finally:
        db.close()


@dashboard_bp.route("/api/crew/certifications/<int:cert_id>/delete", methods=["POST"])
def api_delete_certification(cert_id):
    """Soft-delete a certification record."""
    db = get_db()
    try:
        cert = db.execute("SELECT id FROM certifications WHERE id = ? AND is_active = 1", (cert_id,)).fetchone()
        if not cert:
            return jsonify({"error": "Certification not found"}), 404
        db.execute("UPDATE certifications SET is_active = 0, updated_at = datetime('now') WHERE id = ?", (cert_id,))
        db.commit()
        return jsonify({"status": "deleted"})
    finally:
        db.close()


# ── Projects Page ───────────────────────────────────────


@dashboard_bp.route("/projects")
def projects_page():
    """Dedicated project management page."""
    db = get_db()
    try:
        projects = db.execute("""
            SELECT p.*,
                   (SELECT COUNT(*) FROM receipts r WHERE r.project_id = p.id) as receipt_count,
                   (SELECT COALESCE(SUM(r.total), 0) FROM receipts r WHERE r.project_id = p.id) as total_spend
            FROM projects p ORDER BY p.name
        """).fetchall()
        return _render_module(
            "projects.html", "crewledger", "projects",
            projects=[dict(p) for p in projects],
        )
    finally:
        db.close()


# ── Email Settings ──────────────────────────────────────


@dashboard_bp.route("/settings")
def settings_page():
    """Settings page — email config, links to employee/project management."""
    db = get_db()
    try:
        rows = db.execute("SELECT key, value FROM email_settings").fetchall()
        settings = {r["key"]: r["value"] for r in rows}
        employees = db.execute("SELECT id, first_name FROM employees ORDER BY first_name").fetchall()
        projects = db.execute("SELECT id, name FROM projects WHERE status = 'active' ORDER BY name").fetchall()
        return _render_module(
            "settings.html", "crewledger", "settings",
            settings=settings,
            employees=[dict(e) for e in employees],
            projects=[dict(p) for p in projects],
        )
    finally:
        db.close()


@dashboard_bp.route("/api/settings", methods=["GET"])
def api_get_settings():
    """Get all email settings."""
    db = get_db()
    try:
        rows = db.execute("SELECT key, value FROM email_settings").fetchall()
        return jsonify({r["key"]: r["value"] for r in rows})
    finally:
        db.close()


@dashboard_bp.route("/api/settings", methods=["PUT"])
def api_update_settings():
    """Update email settings."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    allowed_keys = {
        "recipient_email", "frequency", "day_of_week", "time_of_day",
        "include_scope", "include_filter", "enabled",
    }

    db = get_db()
    try:
        for key, value in data.items():
            if key in allowed_keys:
                db.execute(
                    "INSERT OR REPLACE INTO email_settings (key, value, updated_at) VALUES (?, ?, datetime('now'))",
                    (key, str(value)),
                )
        db.commit()
        return jsonify({"status": "updated"})
    finally:
        db.close()


@dashboard_bp.route("/api/settings/send-now", methods=["POST"])
def api_send_report_now():
    """Trigger an immediate email report with current settings."""
    db = get_db()
    try:
        rows = db.execute("SELECT key, value FROM email_settings").fetchall()
        settings = {r["key"]: r["value"] for r in rows}
        recipient = settings.get("recipient_email", "")
        if not recipient:
            return jsonify({"error": "No recipient email configured"}), 400

        # Trigger the existing weekly report send endpoint
        from flask import current_app
        with current_app.test_client() as client:
            resp = client.post(f"/reports/weekly/send?recipient={recipient}")
            if resp.status_code == 200:
                return jsonify({"status": "sent", "recipient": recipient})
            return jsonify({"error": "Failed to send report"}), 500
    finally:
        db.close()


# ── Dashboard Summary API (week-over-week, breakdowns) ───────


@dashboard_bp.route("/api/dashboard/summary", methods=["GET"])
def dashboard_summary():
    """Home screen data: week total, quick stats, flagged count, recent activity.

    Query params:
        week_start — YYYY-MM-DD (default: last Monday)
        week_end   — YYYY-MM-DD (default: last Sunday)
    """
    week_start = request.args.get("week_start")
    week_end = request.args.get("week_end")

    if not week_start or not week_end:
        week_start, week_end = _default_week_range()

    ws = datetime.strptime(week_start, "%Y-%m-%d").date()
    we = datetime.strptime(week_end, "%Y-%m-%d").date()
    prev_start = (ws - timedelta(days=7)).isoformat()
    prev_end = (we - timedelta(days=7)).isoformat()

    db = get_db()
    try:
        current = db.execute(
            """SELECT COALESCE(SUM(total), 0) AS total_spend, COUNT(*) AS receipt_count
               FROM receipts
               WHERE purchase_date >= ? AND purchase_date <= ?
                 AND status IN ('confirmed', 'pending')""",
            (week_start, week_end),
        ).fetchone()

        previous = db.execute(
            """SELECT COALESCE(SUM(total), 0) AS total_spend, COUNT(*) AS receipt_count
               FROM receipts
               WHERE purchase_date >= ? AND purchase_date <= ?
                 AND status IN ('confirmed', 'pending')""",
            (prev_start, prev_end),
        ).fetchone()

        flagged = db.execute(
            "SELECT COUNT(*) AS cnt FROM receipts WHERE status = 'flagged'"
        ).fetchone()

        by_crew = db.execute(
            """SELECT e.id AS employee_id, e.first_name, e.full_name, e.crew,
                      COALESCE(SUM(r.total), 0) AS spend, COUNT(r.id) AS receipt_count
               FROM receipts r JOIN employees e ON r.employee_id = e.id
               WHERE r.purchase_date >= ? AND r.purchase_date <= ?
                 AND r.status IN ('confirmed', 'pending')
               GROUP BY e.id ORDER BY spend DESC""",
            (week_start, week_end),
        ).fetchall()

        by_project = db.execute(
            """SELECT COALESCE(p.name, r.matched_project_name, 'Unassigned') AS project_name,
                      COALESCE(SUM(r.total), 0) AS spend, COUNT(r.id) AS receipt_count
               FROM receipts r LEFT JOIN projects p ON r.project_id = p.id
               WHERE r.purchase_date >= ? AND r.purchase_date <= ?
                 AND r.status IN ('confirmed', 'pending')
               GROUP BY project_name ORDER BY spend DESC""",
            (week_start, week_end),
        ).fetchall()

        recent = db.execute(
            """SELECT r.id, r.vendor_name, r.total, r.purchase_date, r.status,
                      r.matched_project_name, r.created_at, r.image_path,
                      e.id AS employee_id, e.first_name, e.full_name,
                      p.name AS project_name
               FROM receipts r JOIN employees e ON r.employee_id = e.id
               LEFT JOIN projects p ON r.project_id = p.id
               ORDER BY r.created_at DESC LIMIT 10""",
        ).fetchall()

        return jsonify({
            "week_start": week_start,
            "week_end": week_end,
            "current_week": {"total_spend": round(current["total_spend"], 2), "receipt_count": current["receipt_count"]},
            "previous_week": {"total_spend": round(previous["total_spend"], 2), "receipt_count": previous["receipt_count"]},
            "flagged_count": flagged["cnt"],
            "by_crew": [{"id": r["employee_id"], "name": r["full_name"] or r["first_name"], "crew": r["crew"] or "", "spend": round(r["spend"], 2), "receipt_count": r["receipt_count"]} for r in by_crew],
            "by_project": [{"name": r["project_name"], "spend": round(r["spend"], 2), "receipt_count": r["receipt_count"]} for r in by_project],
            "recent_activity": [{"id": r["id"], "vendor": r["vendor_name"] or "Unknown", "total": r["total"], "date": r["purchase_date"], "status": r["status"], "project": r["project_name"] or r["matched_project_name"] or "", "employee": r["full_name"] or r["first_name"], "employee_id": r["employee_id"], "has_image": bool(r["image_path"]), "created_at": r["created_at"]} for r in recent],
        })
    finally:
        db.close()


# ── Flagged Receipt Review Queue ─────────────────────────────


@dashboard_bp.route("/api/dashboard/flagged", methods=["GET"])
def flagged_receipts():
    """Return all flagged receipts for the review queue."""
    db = get_db()
    try:
        rows = db.execute(
            """SELECT r.id, r.vendor_name, r.total, r.purchase_date, r.status,
                      r.flag_reason, r.image_path, r.is_missed_receipt, r.is_return,
                      r.matched_project_name, r.created_at, r.subtotal, r.tax,
                      r.payment_method,
                      e.first_name, e.full_name, p.name AS project_name
               FROM receipts r
               JOIN employees e ON r.employee_id = e.id
               LEFT JOIN projects p ON r.project_id = p.id
               WHERE r.status = 'flagged'
               ORDER BY r.created_at DESC""",
        ).fetchall()

        results = []
        for r in rows:
            items = db.execute(
                "SELECT item_name, quantity, unit_price, extended_price FROM line_items WHERE receipt_id = ? ORDER BY id",
                (r["id"],),
            ).fetchall()
            results.append({
                "id": r["id"], "vendor": r["vendor_name"] or "Unknown", "total": r["total"],
                "subtotal": r["subtotal"], "tax": r["tax"], "date": r["purchase_date"],
                "flag_reason": r["flag_reason"] or "No reason specified",
                "image_path": r["image_path"], "is_missed": bool(r["is_missed_receipt"]),
                "is_return": bool(r["is_return"]), "payment_method": r["payment_method"] or "",
                "project": r["project_name"] or r["matched_project_name"] or "",
                "employee": r["full_name"] or r["first_name"], "created_at": r["created_at"],
                "line_items": [{"name": i["item_name"], "qty": i["quantity"], "price": i["extended_price"]} for i in items],
            })
        return jsonify({"flagged": results, "count": len(results)})
    finally:
        db.close()


@dashboard_bp.route("/api/dashboard/flagged/<int:receipt_id>/approve", methods=["POST"])
def approve_receipt(receipt_id):
    """Approve a flagged receipt — sets status to confirmed."""
    db = get_db()
    try:
        receipt = db.execute("SELECT id, status FROM receipts WHERE id = ?", (receipt_id,)).fetchone()
        if not receipt:
            return jsonify({"error": "Receipt not found"}), 404
        if receipt["status"] != "flagged":
            return jsonify({"error": "Receipt is not flagged"}), 400
        db.execute("UPDATE receipts SET status = 'confirmed', confirmed_at = datetime('now') WHERE id = ?", (receipt_id,))
        db.commit()
        log.info("Receipt #%d approved via dashboard", receipt_id)
        return jsonify({"status": "approved", "id": receipt_id})
    finally:
        db.close()


@dashboard_bp.route("/api/dashboard/flagged/<int:receipt_id>/dismiss", methods=["POST"])
def dismiss_receipt(receipt_id):
    """Dismiss a flagged receipt — sets status to rejected."""
    db = get_db()
    try:
        receipt = db.execute("SELECT id, status FROM receipts WHERE id = ?", (receipt_id,)).fetchone()
        if not receipt:
            return jsonify({"error": "Receipt not found"}), 404
        if receipt["status"] != "flagged":
            return jsonify({"error": "Receipt is not flagged"}), 400
        db.execute("UPDATE receipts SET status = 'rejected' WHERE id = ?", (receipt_id,))
        db.commit()
        log.info("Receipt #%d dismissed via dashboard", receipt_id)
        return jsonify({"status": "dismissed", "id": receipt_id})
    finally:
        db.close()


@dashboard_bp.route("/api/dashboard/flagged/<int:receipt_id>/edit", methods=["POST"])
def edit_receipt(receipt_id):
    """Edit a flagged receipt's fields, then approve it."""
    if not check_permission(None, "crewledger", "edit"):
        return jsonify({"error": "Insufficient permissions"}), 403
    data = request.get_json(silent=True) or {}
    db = get_db()
    try:
        receipt = db.execute("SELECT * FROM receipts WHERE id = ?", (receipt_id,)).fetchone()
        if not receipt:
            return jsonify({"error": "Receipt not found"}), 404
        updatable = {
            "vendor_name": data.get("vendor"), "total": data.get("total"),
            "subtotal": data.get("subtotal"), "tax": data.get("tax"),
            "purchase_date": data.get("date"), "payment_method": data.get("payment_method"),
            "matched_project_name": data.get("project"),
        }
        updates = {k: v for k, v in updatable.items() if v is not None}
        # Log audit trail for each changed field
        for field, new_val in updates.items():
            old_val = receipt[field]
            if str(old_val) != str(new_val):
                db.execute(
                    "INSERT INTO receipt_edits (receipt_id, field_changed, old_value, new_value, edited_by) VALUES (?, ?, ?, ?, ?)",
                    (receipt_id, field, str(old_val) if old_val is not None else None, str(new_val), "dashboard"),
                )
        if updates:
            set_clause = ", ".join(f"{k} = ?" for k in updates)
            values = list(updates.values()) + [receipt_id]
            db.execute(f"UPDATE receipts SET {set_clause}, status = 'confirmed', confirmed_at = datetime('now') WHERE id = ?", values)
        else:
            db.execute("UPDATE receipts SET status = 'confirmed', confirmed_at = datetime('now') WHERE id = ?", (receipt_id,))
        db.commit()
        log.info("Receipt #%d edited and approved via dashboard", receipt_id)
        return jsonify({"status": "updated", "id": receipt_id})
    finally:
        db.close()


# ── Receipt Editing (General — with audit trail) ────────────


@dashboard_bp.route("/api/receipts/<int:receipt_id>/edit", methods=["POST"])
def api_edit_receipt(receipt_id):
    """Edit any receipt's fields with full audit trail.

    Accepts JSON with any of: vendor_name, vendor_city, vendor_state,
    purchase_date, subtotal, tax, total, payment_method, notes,
    matched_project_name, project_id, employee_id.
    """
    if not check_permission(None, "crewledger", "edit"):
        return jsonify({"error": "Insufficient permissions"}), 403
    data = request.get_json(silent=True) or {}
    if not data:
        return jsonify({"error": "No data provided"}), 400

    db = get_db()
    try:
        receipt = db.execute("SELECT * FROM receipts WHERE id = ?", (receipt_id,)).fetchone()
        if not receipt:
            return jsonify({"error": "Receipt not found"}), 404

        allowed_fields = {
            "vendor_name", "vendor_city", "vendor_state", "purchase_date",
            "subtotal", "tax", "total", "payment_method", "notes",
            "matched_project_name", "project_id", "category_id", "status", "duplicate_of",
            "employee_id",
        }
        updates = {k: v for k, v in data.items() if k in allowed_fields}
        if not updates:
            return jsonify({"error": "No valid fields to update"}), 400

        # Log each change to audit trail
        for field, new_val in updates.items():
            old_val = receipt[field]
            if str(old_val) != str(new_val):
                # For employee_id changes, log human-readable names
                if field == "employee_id":
                    old_emp = db.execute("SELECT first_name, full_name FROM employees WHERE id = ?", (old_val,)).fetchone()
                    new_emp = db.execute("SELECT first_name, full_name FROM employees WHERE id = ?", (new_val,)).fetchone()
                    old_display = (old_emp["full_name"] or old_emp["first_name"]) if old_emp else str(old_val)
                    new_display = (new_emp["full_name"] or new_emp["first_name"]) if new_emp else str(new_val)
                    db.execute(
                        "INSERT INTO receipt_edits (receipt_id, field_changed, old_value, new_value, edited_by) VALUES (?, ?, ?, ?, ?)",
                        (receipt_id, "employee_id", old_display, new_display, "dashboard"),
                    )
                elif field == "project_id":
                    old_proj = db.execute("SELECT name FROM projects WHERE id = ?", (old_val,)).fetchone() if old_val else None
                    new_proj = db.execute("SELECT name FROM projects WHERE id = ?", (new_val,)).fetchone() if new_val else None
                    old_display = old_proj["name"] if old_proj else None
                    new_display = new_proj["name"] if new_proj else None
                    db.execute(
                        "INSERT INTO receipt_edits (receipt_id, field_changed, old_value, new_value, edited_by) VALUES (?, ?, ?, ?, ?)",
                        (receipt_id, "project", old_display, new_display, "dashboard"),
                    )
                else:
                    db.execute(
                        "INSERT INTO receipt_edits (receipt_id, field_changed, old_value, new_value, edited_by) VALUES (?, ?, ?, ?, ?)",
                        (receipt_id, field, str(old_val) if old_val is not None else None, str(new_val), "dashboard"),
                    )

        set_clause = ", ".join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [receipt_id]
        db.execute(f"UPDATE receipts SET {set_clause} WHERE id = ?", values)

        # Clear conversation state if status changed via dashboard
        if "status" in updates and updates["status"] in ("confirmed", "deleted"):
            employee_id = receipt["employee_id"]
            convo = db.execute(
                "SELECT id FROM conversation_state WHERE employee_id = ? AND receipt_id = ?",
                (employee_id, receipt_id),
            ).fetchone()
            if convo:
                db.execute(
                    "UPDATE conversation_state SET state = 'idle', updated_at = datetime('now') WHERE id = ?",
                    (convo["id"],),
                )

        db.commit()

        log.info("Receipt #%d edited via dashboard (%s)", receipt_id, ", ".join(updates.keys()))
        return jsonify({"status": "updated", "id": receipt_id, "fields_changed": list(updates.keys())})
    finally:
        db.close()


@dashboard_bp.route("/api/receipts/<int:receipt_id>/edits", methods=["GET"])
def api_receipt_edit_history(receipt_id):
    """Get the audit trail for a receipt."""
    db = get_db()
    try:
        receipt = db.execute("SELECT id FROM receipts WHERE id = ?", (receipt_id,)).fetchone()
        if not receipt:
            return jsonify({"error": "Receipt not found"}), 404

        edits = db.execute(
            "SELECT * FROM receipt_edits WHERE receipt_id = ? ORDER BY edited_at DESC",
            (receipt_id,),
        ).fetchall()
        return jsonify({"receipt_id": receipt_id, "edits": [dict(e) for e in edits]})
    finally:
        db.close()


@dashboard_bp.route("/api/receipts/<int:receipt_id>/delete", methods=["POST"])
def api_delete_receipt(receipt_id):
    """Soft-delete a receipt (set status to 'deleted')."""
    if not check_permission(None, "crewledger", "edit"):
        return jsonify({"error": "Insufficient permissions"}), 403
    db = get_db()
    try:
        receipt = db.execute("SELECT * FROM receipts WHERE id = ?", (receipt_id,)).fetchone()
        if not receipt:
            return jsonify({"error": "Receipt not found"}), 404

        old_status = receipt["status"]
        db.execute("UPDATE receipts SET status = 'deleted' WHERE id = ?", (receipt_id,))
        db.execute(
            "INSERT INTO receipt_edits (receipt_id, field_changed, old_value, new_value, edited_by) VALUES (?, 'status', ?, 'deleted', 'management')",
            (receipt_id, old_status),
        )
        # Clear conversation state so employee can submit new receipts
        employee_id = receipt["employee_id"]
        convo = db.execute(
            "SELECT id FROM conversation_state WHERE employee_id = ? AND receipt_id = ?",
            (employee_id, receipt_id),
        ).fetchone()
        if convo:
            db.execute(
                "UPDATE conversation_state SET state = 'idle', updated_at = datetime('now') WHERE id = ?",
                (convo["id"],),
            )
        db.commit()
        log.info("Receipt #%d soft-deleted (was %s)", receipt_id, old_status)
        return jsonify({"status": "deleted", "id": receipt_id})
    finally:
        db.close()


@dashboard_bp.route("/api/receipts/<int:receipt_id>/restore", methods=["POST"])
def api_restore_receipt(receipt_id):
    """Restore a deleted or duplicate receipt back to confirmed."""
    db = get_db()
    try:
        receipt = db.execute("SELECT * FROM receipts WHERE id = ?", (receipt_id,)).fetchone()
        if not receipt:
            return jsonify({"error": "Receipt not found"}), 404

        old_status = receipt["status"]
        db.execute("UPDATE receipts SET status = 'confirmed', duplicate_of = NULL WHERE id = ?", (receipt_id,))
        db.execute(
            "INSERT INTO receipt_edits (receipt_id, field_changed, old_value, new_value, edited_by) VALUES (?, 'status', ?, 'confirmed', 'management')",
            (receipt_id, old_status),
        )
        db.commit()
        log.info("Receipt #%d restored to confirmed (was %s)", receipt_id, old_status)
        return jsonify({"status": "restored", "id": receipt_id})
    finally:
        db.close()


@dashboard_bp.route("/api/receipts/<int:receipt_id>/duplicate", methods=["POST"])
def api_mark_duplicate(receipt_id):
    """Mark a receipt as duplicate of another."""
    data = request.get_json(silent=True) or {}
    duplicate_of = data.get("duplicate_of")

    db = get_db()
    try:
        receipt = db.execute("SELECT * FROM receipts WHERE id = ?", (receipt_id,)).fetchone()
        if not receipt:
            return jsonify({"error": "Receipt not found"}), 404

        if duplicate_of:
            original = db.execute("SELECT id FROM receipts WHERE id = ?", (duplicate_of,)).fetchone()
            if not original:
                return jsonify({"error": "Original receipt not found"}), 404

        old_status = receipt["status"]
        db.execute("UPDATE receipts SET status = 'duplicate', duplicate_of = ? WHERE id = ?", (duplicate_of, receipt_id))
        db.execute(
            "INSERT INTO receipt_edits (receipt_id, field_changed, old_value, new_value, edited_by) VALUES (?, 'status', ?, 'duplicate', 'management')",
            (receipt_id, old_status),
        )
        db.commit()
        log.info("Receipt #%d marked as duplicate of #%s", receipt_id, duplicate_of)
        return jsonify({"status": "duplicate", "id": receipt_id, "duplicate_of": duplicate_of})
    finally:
        db.close()


@dashboard_bp.route("/api/receipts/<int:receipt_id>/notes", methods=["PUT"])
def api_update_receipt_notes(receipt_id):
    """Update the notes field on a receipt."""
    data = request.get_json(silent=True) or {}
    notes = data.get("notes", "")

    db = get_db()
    try:
        receipt = db.execute("SELECT id, notes FROM receipts WHERE id = ?", (receipt_id,)).fetchone()
        if not receipt:
            return jsonify({"error": "Receipt not found"}), 404

        old_notes = receipt["notes"]
        if old_notes != notes:
            db.execute(
                "INSERT INTO receipt_edits (receipt_id, field_changed, old_value, new_value, edited_by) VALUES (?, ?, ?, ?, ?)",
                (receipt_id, "notes", old_notes, notes, "dashboard"),
            )

        db.execute("UPDATE receipts SET notes = ? WHERE id = ?", (notes, receipt_id))
        db.commit()
        return jsonify({"status": "updated", "id": receipt_id})
    finally:
        db.close()


# ── Search & Filter (paginated) ──────────────────────────────


@dashboard_bp.route("/api/dashboard/search", methods=["GET"])
def search_receipts():
    """Search receipts with filters and pagination."""
    date_start = request.args.get("date_start")
    date_end = request.args.get("date_end")
    employee = request.args.get("employee")
    employee_id = request.args.get("employee_id", type=int)
    project = request.args.get("project")
    vendor = request.args.get("vendor")
    category = request.args.get("category")
    amount_min = request.args.get("amount_min", type=float)
    amount_max = request.args.get("amount_max", type=float)
    status = request.args.get("status")
    sort_by = request.args.get("sort", "date")
    order = request.args.get("order", "desc")
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 25, type=int)

    db = get_db()
    try:
        sql = """SELECT r.id, r.vendor_name, r.vendor_city, r.vendor_state,
                        r.total, r.subtotal, r.tax, r.purchase_date, r.status,
                        r.payment_method, r.image_path, r.flag_reason,
                        r.is_missed_receipt, r.is_return, r.matched_project_name,
                        r.created_at, e.first_name, e.full_name, e.id AS employee_id,
                        p.name AS project_name
                 FROM receipts r
                 JOIN employees e ON r.employee_id = e.id
                 LEFT JOIN projects p ON r.project_id = p.id
                 WHERE 1=1"""
        params: list = []

        if date_start:
            sql += " AND r.purchase_date >= ?"
            params.append(date_start)
        if date_end:
            sql += " AND r.purchase_date <= ?"
            params.append(date_end)
        if employee:
            sql += " AND (e.first_name LIKE ? OR e.full_name LIKE ?)"
            params.extend([f"%{employee}%", f"%{employee}%"])
        if employee_id is not None:
            sql += " AND e.id = ?"
            params.append(employee_id)
        if project:
            sql += " AND (p.name LIKE ? OR r.matched_project_name LIKE ?)"
            params.extend([f"%{project}%", f"%{project}%"])
        if vendor:
            sql += " AND r.vendor_name LIKE ?"
            params.append(f"%{vendor}%")
        if amount_min is not None:
            sql += " AND r.total >= ?"
            params.append(amount_min)
        if amount_max is not None:
            sql += " AND r.total <= ?"
            params.append(amount_max)
        if status:
            sql += " AND r.status = ?"
            params.append(status)
        if category:
            sql += " AND r.id IN (SELECT li.receipt_id FROM line_items li JOIN categories c ON li.category_id = c.id WHERE c.name LIKE ?)"
            params.append(f"%{category}%")

        sort_map = {"date": "r.purchase_date", "amount": "r.total", "employee": "e.first_name", "vendor": "r.vendor_name", "project": "COALESCE(p.name, r.matched_project_name)"}
        sort_col = sort_map.get(sort_by, "r.purchase_date")
        sort_dir = "ASC" if order == "asc" else "DESC"
        sql += f" ORDER BY {sort_col} {sort_dir}"

        count_sql = f"SELECT COUNT(*) AS cnt FROM ({sql})"
        total_count = db.execute(count_sql, params).fetchone()["cnt"]

        offset = (page - 1) * per_page
        sql += " LIMIT ? OFFSET ?"
        params.extend([per_page, offset])
        rows = db.execute(sql, params).fetchall()

        results = []
        for r in rows:
            items = db.execute(
                "SELECT li.item_name, li.quantity, li.extended_price, c.name AS category_name FROM line_items li LEFT JOIN categories c ON li.category_id = c.id WHERE li.receipt_id = ? ORDER BY li.id",
                (r["id"],),
            ).fetchall()
            results.append({
                "id": r["id"], "vendor": r["vendor_name"] or "Unknown",
                "total": r["total"], "date": r["purchase_date"], "status": r["status"],
                "payment_method": r["payment_method"] or "", "image_path": r["image_path"],
                "project": r["project_name"] or r["matched_project_name"] or "",
                "employee": r["full_name"] or r["first_name"], "employee_id": r["employee_id"],
                "created_at": r["created_at"],
                "line_items": [{"name": i["item_name"], "qty": i["quantity"], "price": i["extended_price"], "category": i["category_name"]} for i in items],
            })

        return jsonify({"results": results, "total": total_count, "page": page, "per_page": per_page, "total_pages": max(1, -(-total_count // per_page))})
    finally:
        db.close()


# ── Employee Receipts Drill-down ─────────────────────────────


@dashboard_bp.route("/api/dashboard/employee/<int:employee_id>/receipts", methods=["GET"])
def employee_receipts(employee_id):
    """Return all receipts for a given employee."""
    status_filter = request.args.get("status")
    limit = request.args.get("limit", 50, type=int)

    db = get_db()
    try:
        emp = db.execute("SELECT id, first_name, full_name, phone_number, crew FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not emp:
            return jsonify({"error": "Employee not found"}), 404

        sql = """SELECT r.id, r.vendor_name, r.total, r.subtotal, r.tax,
                        r.purchase_date, r.status, r.payment_method,
                        r.image_path, r.flag_reason, r.is_missed_receipt,
                        r.is_return, r.matched_project_name, r.created_at,
                        p.name AS project_name
                 FROM receipts r LEFT JOIN projects p ON r.project_id = p.id
                 WHERE r.employee_id = ?"""
        params: list = [employee_id]
        if status_filter:
            sql += " AND r.status = ?"
            params.append(status_filter)
        sql += " ORDER BY r.created_at DESC LIMIT ?"
        params.append(limit)

        rows = db.execute(sql, params).fetchall()
        results = []
        for r in rows:
            items = db.execute("SELECT item_name, quantity, unit_price, extended_price FROM line_items WHERE receipt_id = ? ORDER BY id", (r["id"],)).fetchall()
            results.append({
                "id": r["id"], "vendor": r["vendor_name"] or "Unknown", "total": r["total"],
                "date": r["purchase_date"], "status": r["status"],
                "project": r["project_name"] or r["matched_project_name"] or "",
                "created_at": r["created_at"],
                "line_items": [{"name": i["item_name"], "qty": i["quantity"], "price": i["extended_price"]} for i in items],
            })

        return jsonify({
            "employee": {"id": emp["id"], "name": emp["full_name"] or emp["first_name"], "phone": emp["phone_number"], "crew": emp["crew"] or ""},
            "receipts": results, "count": len(results),
        })
    finally:
        db.close()


# ── Export Helpers ────────────────────────────────────────────


def _export_csv(receipts: list) -> Response:
    """Export as standard CSV (Google Sheets compatible)."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Employee", "Vendor", "Project", "Subtotal", "Tax", "Total", "Payment Method", "Status", "Notes"])
    for r in receipts:
        writer.writerow([
            r.get("purchase_date", ""),
            r.get("employee_name", ""),
            r.get("vendor_name", ""),
            r.get("project_name") or r.get("matched_project_name", ""),
            r.get("subtotal", ""),
            r.get("tax", ""),
            r.get("total", ""),
            r.get("payment_method", ""),
            r.get("status", ""),
            r.get("notes", ""),
        ])

    resp = Response(output.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = f"attachment; filename=crewledger_export_{datetime.now().strftime('%Y%m%d')}.csv"
    return resp


def _export_quickbooks_csv(receipts: list) -> Response:
    """Export as QuickBooks IIF/CSV format for expense import."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Vendor", "Account", "Amount", "Memo", "Payment Method"])
    for r in receipts:
        project = r.get("project_name") or r.get("matched_project_name", "")
        notes = r.get("notes", "")
        memo = f"Employee: {r.get('employee_name', '')} | Project: {project}"
        if notes:
            memo += f" | Notes: {notes}"
        writer.writerow([
            r.get("purchase_date", ""),
            r.get("vendor_name", ""),
            "Materials & Supplies",
            r.get("total", ""),
            memo,
            r.get("payment_method", ""),
        ])

    resp = Response(output.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = f"attachment; filename=crewledger_quickbooks_{datetime.now().strftime('%Y%m%d')}.csv"
    return resp


def _export_excel(receipts: list) -> Response:
    """Export as Excel (.xlsx) with formatting."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CrewLedger Export"

    # Header row
    headers = ["Date", "Employee", "Vendor", "Project", "Subtotal", "Tax", "Total", "Payment Method", "Status", "Notes"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, r in enumerate(receipts, 2):
        ws.cell(row=row_idx, column=1, value=r.get("purchase_date", ""))
        ws.cell(row=row_idx, column=2, value=r.get("employee_name", ""))
        ws.cell(row=row_idx, column=3, value=r.get("vendor_name", ""))
        ws.cell(row=row_idx, column=4, value=r.get("project_name") or r.get("matched_project_name", ""))
        ws.cell(row=row_idx, column=5, value=r.get("subtotal") or 0).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=6, value=r.get("tax") or 0).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=7, value=r.get("total") or 0).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=8, value=r.get("payment_method", ""))
        ws.cell(row=row_idx, column=9, value=r.get("status", ""))
        ws.cell(row=row_idx, column=10, value=r.get("notes", ""))

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # Total row
    total_row = len(receipts) + 2
    ws.cell(row=total_row, column=6, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=sum(r.get("total", 0) or 0 for r in receipts)).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"crewledger_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
    )


# ── Data Helpers ─────────────────────────────────────────────


def _get_dashboard_stats(db) -> dict:
    """Summary stats for the dashboard home screen."""
    row = db.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN created_at >= date('now', 'weekday 1', '-7 days') THEN total ELSE 0 END), 0) as week_spend,
            COALESCE(SUM(CASE WHEN created_at >= date('now', 'start of month') THEN total ELSE 0 END), 0) as month_spend,
            COUNT(*) as total_receipts,
            COALESCE(SUM(CASE WHEN status = 'flagged' THEN 1 ELSE 0 END), 0) as flagged_count,
            COALESCE(SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END), 0) as pending_count,
            COALESCE(SUM(CASE WHEN status = 'confirmed' THEN 1 ELSE 0 END), 0) as confirmed_count
        FROM receipts
        WHERE status NOT IN ('deleted', 'duplicate')
    """).fetchone()

    employee_count = db.execute("SELECT COUNT(*) as cnt FROM employees WHERE is_active = 1").fetchone()["cnt"]
    project_count = db.execute("SELECT COUNT(*) as cnt FROM projects WHERE status = 'active'").fetchone()["cnt"]

    unknown_count = db.execute("SELECT COUNT(*) as cnt FROM unknown_contacts").fetchone()["cnt"]

    return {
        "week_spend": round(row["week_spend"], 2),
        "month_spend": round(row["month_spend"], 2),
        "total_receipts": row["total_receipts"],
        "flagged_count": row["flagged_count"],
        "pending_count": row["pending_count"],
        "confirmed_count": row["confirmed_count"],
        "employee_count": employee_count,
        "project_count": project_count,
        "unknown_count": unknown_count,
    }


def _get_flagged_receipts(db, limit=20) -> list:
    """Flagged receipts for the review queue."""
    rows = db.execute("""
        SELECT r.*, e.first_name as employee_name, p.name as project_name
        FROM receipts r
        LEFT JOIN employees e ON r.employee_id = e.id
        LEFT JOIN projects p ON r.project_id = p.id
        WHERE r.status = 'flagged'
        ORDER BY r.created_at DESC
        LIMIT ?
    """, (limit,)).fetchall()
    return [_row_to_dict(r) for r in rows]


def _get_recent_receipts(db, limit=10) -> list:
    """Most recent receipts for the activity feed."""
    rows = db.execute("""
        SELECT r.*, e.first_name as employee_name, p.name as project_name
        FROM receipts r
        LEFT JOIN employees e ON r.employee_id = e.id
        LEFT JOIN projects p ON r.project_id = p.id
        ORDER BY r.created_at DESC
        LIMIT ?
    """, (limit,)).fetchall()
    return [_row_to_dict(r) for r in rows]


def _query_receipts(db, args) -> list:
    """Query receipts with filters and sorting."""
    conditions = []
    params = []

    # Period filter — use purchase_date (actual receipt date), fall back to created_at
    date_col = "COALESCE(r.purchase_date, date(r.created_at))"
    period = args.get("period", "all")
    if period == "today":
        conditions.append(f"{date_col} >= date('now')")
    elif period == "week":
        conditions.append(f"{date_col} >= date('now', '-' || ((strftime('%w', 'now') + 6) % 7) || ' days')")
    elif period == "month":
        conditions.append(f"{date_col} >= date('now', 'start of month')")
    elif period == "ytd":
        conditions.append(f"{date_col} >= date('now', 'start of year')")

    # Custom date range
    start = args.get("start")
    end = args.get("end")
    if start:
        conditions.append(f"{date_col} >= ?")
        params.append(start)
    if end:
        conditions.append(f"{date_col} <= ?")
        params.append(end)

    # Filters
    employee_id = args.get("employee")
    if employee_id:
        conditions.append("r.employee_id = ?")
        params.append(employee_id)

    project_id = args.get("project")
    if project_id:
        conditions.append("r.project_id = ?")
        params.append(project_id)

    vendor = args.get("vendor")
    if vendor:
        conditions.append("r.vendor_name LIKE ?")
        params.append(f"%{vendor}%")

    status = args.get("status")
    if status:
        conditions.append("r.status = ?")
        params.append(status)

    # Exclude deleted and duplicate receipts by default
    include_hidden = args.get("include_hidden", "0")
    if include_hidden != "1" and not status:
        conditions.append("r.status NOT IN ('deleted', 'duplicate')")

    where = " AND ".join(conditions) if conditions else "1=1"

    # Sorting
    sort_map = {
        "date": "COALESCE(r.purchase_date, date(r.created_at))",
        "employee": "e.first_name",
        "vendor": "r.vendor_name",
        "project": "p.name",
        "amount": "r.total",
        "status": "r.status",
    }
    sort_col = sort_map.get(args.get("sort", "date"), "COALESCE(r.purchase_date, date(r.created_at))")
    order = "ASC" if args.get("order") == "asc" else "DESC"

    rows = db.execute(f"""
        SELECT r.*, e.first_name as employee_name, e.crew,
               p.name as project_name,
               cat.name as category
        FROM receipts r
        LEFT JOIN employees e ON r.employee_id = e.id
        LEFT JOIN projects p ON r.project_id = p.id
        LEFT JOIN categories cat ON r.category_id = cat.id
        WHERE {where}
        ORDER BY {sort_col} {order}
        LIMIT 500
    """, params).fetchall()

    return [_row_to_dict(r) for r in rows]


def _get_receipt_detail(db, receipt_id: int) -> dict | None:
    """Single receipt with line items."""
    row = db.execute("""
        SELECT r.*, e.first_name as employee_name, e.crew,
               p.name as project_name,
               cat.name as category
        FROM receipts r
        LEFT JOIN employees e ON r.employee_id = e.id
        LEFT JOIN projects p ON r.project_id = p.id
        LEFT JOIN categories cat ON r.category_id = cat.id
        WHERE r.id = ?
    """, (receipt_id,)).fetchone()

    if not row:
        return None

    result = _row_to_dict(row)

    items = db.execute("""
        SELECT li.*, c.name as category_name
        FROM line_items li
        LEFT JOIN categories c ON li.category_id = c.id
        WHERE li.receipt_id = ?
        ORDER BY li.id
    """, (receipt_id,)).fetchall()

    result["line_items"] = [dict(i) for i in items]
    return result


def _get_unknown_contacts(db, limit=10) -> list:
    """Recent unknown contact attempts for dashboard."""
    rows = db.execute("""
        SELECT * FROM unknown_contacts ORDER BY created_at DESC LIMIT ?
    """, (limit,)).fetchall()
    return [dict(r) for r in rows]


def _default_week_range() -> tuple[str, str]:
    """Return (last Monday, last Sunday) as YYYY-MM-DD strings."""
    today = datetime.now().date()
    days_since_monday = today.weekday()
    if days_since_monday == 0:
        last_monday = today - timedelta(days=7)
    else:
        last_monday = today - timedelta(days=days_since_monday + 7)
    last_sunday = last_monday + timedelta(days=6)
    return last_monday.isoformat(), last_sunday.isoformat()


def _row_to_dict(row) -> dict:
    """Convert a sqlite3.Row to a plain dict."""
    d = dict(row)
    # Add image URL if image exists
    if d.get("image_path"):
        filename = Path(d["image_path"]).name
        d["image_url"] = f"/receipts/image/{filename}"
    else:
        d["image_url"] = None
    # Ensure notes is always present (may be None)
    if "notes" not in d:
        d["notes"] = None
    return d
